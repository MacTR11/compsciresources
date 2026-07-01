#!/usr/bin/env python3
"""
Build a Personal Learning Checklist (PLC) spreadsheet: every "I can..."
objective for all 24 subtopics (from the lesson deck-data), with RAG
dropdowns, auto-colour and progress counts — for student self-assessment
and teacher tracking.

Output: Teacher-Toolkit/PLC/Personal-Learning-Checklist.xlsx
"""
import os, glob, json, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(REPO, "source", "teacher-toolkit", "deck-data")
OUT = os.path.join(REPO, "Teacher-Toolkit", "PLC")

HEAD = PatternFill("solid", fgColor="1A365D"); HEADF = Font(bold=True, color="FFFFFF")
SECT = PatternFill("solid", fgColor="2B6CB0"); SECTF = Font(bold=True, color="FFFFFF")
THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def code_key(fn):
    m = re.match(r"(\d+)\.(\d+)\.(\d+)", os.path.basename(fn))
    return tuple(int(x) for x in m.groups()) if m else (9, 9, 9)


def main():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Checklist"
    ws.append(["Spec", "I can…  (tick your confidence)", "RAG", "Date", "Notes / evidence"])
    for c in range(1, 6):
        ws.cell(row=1, column=c).fill = HEAD; ws.cell(row=1, column=c).font = HEADF
        ws.cell(row=1, column=c).alignment = CENTER
    ws.freeze_panes = "A2"

    rag_rows = []
    files = sorted(glob.glob(os.path.join(DATA, "*.json")), key=code_key)
    for f in files:
        d = json.load(open(f, encoding="utf-8"))
        code = re.match(r"(\d+\.\d+\.\d+)", os.path.basename(f)).group(1)
        # section header
        ws.append([code, d.get("subtopic", code).split(" ", 1)[-1], "", "", ""])
        r = ws.max_row
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = SECT; ws.cell(row=r, column=c).font = SECTF
        for obj in d.get("objectives", []):
            ws.append(["", obj, "", "", ""])
            rag_rows.append(ws.max_row)

    widths = {"A": 9, "B": 78, "C": 7, "D": 12, "E": 34}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = THIN; cell.alignment = WRAP
        row[2].alignment = CENTER

    dv = DataValidation(type="list", formula1='"R,A,G"', allow_blank=True)
    dv.prompt = "R = not yet, A = getting there, G = confident"
    ws.add_data_validation(dv)
    for r in rag_rows:
        dv.add(ws.cell(row=r, column=3))
        cellref = f"C{r}"
        ws.conditional_formatting.add(cellref, CellIsRule(operator="equal", formula=['"R"'],
            fill=PatternFill("solid", fgColor="F4CCCC"), font=Font(bold=True, color="990000")))
        ws.conditional_formatting.add(cellref, CellIsRule(operator="equal", formula=['"A"'],
            fill=PatternFill("solid", fgColor="FCE5CD"), font=Font(bold=True, color="996600")))
        ws.conditional_formatting.add(cellref, CellIsRule(operator="equal", formula=['"G"'],
            fill=PatternFill("solid", fgColor="D9EAD3"), font=Font(bold=True, color="006600")))

    # progress counters
    end = ws.max_row + 2
    rng = f"C2:C{ws.max_row}"
    ws.cell(row=end, column=2, value="🔴 Not yet (R):").font = Font(bold=True)
    ws.cell(row=end, column=3, value=f'=COUNTIF({rng},"R")')
    ws.cell(row=end + 1, column=2, value="🟠 Getting there (A):").font = Font(bold=True)
    ws.cell(row=end + 1, column=3, value=f'=COUNTIF({rng},"A")')
    ws.cell(row=end + 2, column=2, value="🟢 Confident (G):").font = Font(bold=True)
    ws.cell(row=end + 2, column=3, value=f'=COUNTIF({rng},"G")')
    ws.cell(row=end + 3, column=2, value="Total objectives:").font = Font(bold=True, italic=True)
    ws.cell(row=end + 3, column=3, value=len(rag_rows))

    os.makedirs(OUT, exist_ok=True)
    wb.save(os.path.join(OUT, "Personal-Learning-Checklist.xlsx"))
    print(f"wrote Personal-Learning-Checklist.xlsx ({len(rag_rows)} objectives)")


if __name__ == "__main__":
    main()
