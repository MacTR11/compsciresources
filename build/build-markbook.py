#!/usr/bin/env python3
"""
Build an interactive teacher MARKBOOK (.xlsx) for H446, tracking the
Key Assessments (KA1-8) and CAP1-5 from the scheme of work with automatic
percentage, grade lookup (editable boundaries) and a class overview.

Output: Teacher-Toolkit/Markbook/H446-Markbook.xlsx
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "Teacher-Toolkit", "Markbook")

HEAD = PatternFill("solid", fgColor="1A365D")
HEADF = Font(bold=True, color="FFFFFF")
SUB = PatternFill("solid", fgColor="2B6CB0")
THIN = Border(*[Side(style="thin", color="CCCCCC")] * 4)
CENTER = Alignment(horizontal="center", vertical="center")

ASSESSMENTS = ["KA1", "KA2", "KA3", "KA4", "KA5", "KA6", "KA7", "KA8",
               "CAP1", "CAP2", "CAP3", "CAP4", "CAP5"]
N_STUDENTS = 24


def grade_sheet(wb):
    ws = wb.create_sheet("Grade Boundaries")
    ws.append(["Min %", "Grade"])
    # ascending for VLOOKUP approximate match; fractions formatted as %
    rows = [(0.0, "U"), (0.40, "E"), (0.50, "D"), (0.60, "C"),
            (0.70, "B"), (0.80, "A"), (0.90, "A*")]
    for pct, g in rows:
        ws.append([pct, g])
    for r in range(2, 2 + len(rows)):
        ws.cell(row=r, column=1).number_format = "0%"
        ws.cell(row=r, column=2).alignment = CENTER
    ws.cell(row=1, column=1).fill = HEAD; ws.cell(row=1, column=1).font = HEADF
    ws.cell(row=1, column=2).fill = HEAD; ws.cell(row=1, column=2).font = HEADF
    ws.column_dimensions["A"].width = 10; ws.column_dimensions["B"].width = 10
    ws["D2"] = "Edit the Min % thresholds to match the grade boundaries you use."
    ws["D2"].font = Font(italic=True, color="666666")
    return ws


def markbook_sheet(wb):
    ws = wb.active; ws.title = "Markbook"
    ncols = 1 + len(ASSESSMENTS) + 2  # Student + assessments + Avg% + Grade
    # Row 1 headers
    ws.cell(row=1, column=1, value="Student")
    for j, a in enumerate(ASSESSMENTS):
        ws.cell(row=1, column=2 + j, value=a)
    ws.cell(row=1, column=2 + len(ASSESSMENTS), value="Avg %")
    ws.cell(row=1, column=3 + len(ASSESSMENTS), value="Grade")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c); cell.fill = HEAD; cell.font = HEADF; cell.alignment = CENTER
    # Row 2: Max mark
    ws.cell(row=2, column=1, value="Max mark →")
    ws.cell(row=2, column=1).font = Font(bold=True, italic=True)
    for j in range(len(ASSESSMENTS)):
        ws.cell(row=2, column=2 + j, value="").fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=2, column=1).fill = PatternFill("solid", fgColor="FFF2CC")
    first_a = get_column_letter(2)
    last_a = get_column_letter(1 + len(ASSESSMENTS))
    avg_col = get_column_letter(2 + len(ASSESSMENTS))
    grade_col = get_column_letter(3 + len(ASSESSMENTS))
    # Student rows
    for i in range(N_STUDENTS):
        r = 3 + i
        ws.cell(row=r, column=1, value=f"Student {i+1}")
        rng = f"{first_a}{r}:{last_a}{r}"
        avg = (f'=IFERROR(SUM({rng})/SUMPRODUCT(({rng}<>"")*${first_a}$2:${last_a}$2),"")')
        ws.cell(row=r, column=2 + len(ASSESSMENTS), value=avg).number_format = "0%"
        ws.cell(row=r, column=3 + len(ASSESSMENTS),
                value=f"=IFERROR(VLOOKUP({avg_col}{r},'Grade Boundaries'!$A$2:$B$8,2,TRUE),\"\")")
        ws.cell(row=r, column=2 + len(ASSESSMENTS)).alignment = CENTER
        ws.cell(row=r, column=3 + len(ASSESSMENTS)).alignment = CENTER
    # widths + borders
    ws.column_dimensions["A"].width = 16
    for j in range(len(ASSESSMENTS)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 7
    ws.column_dimensions[avg_col].width = 9
    ws.column_dimensions[grade_col].width = 9
    for row in ws.iter_rows(min_row=2, max_row=2 + N_STUDENTS, min_col=1, max_col=ncols):
        for cell in row:
            cell.border = THIN
            if cell.column >= 2:
                cell.alignment = CENTER
    ws.freeze_panes = "B3"
    # conditional formatting: Avg% colour scale
    avg_rng = f"{avg_col}3:{avg_col}{2+N_STUDENTS}"
    ws.conditional_formatting.add(avg_rng, ColorScaleRule(
        start_type="num", start_value=0, start_color="F4C7C3",
        mid_type="num", mid_value=0.6, mid_color="FFEB9C",
        end_type="num", end_value=1, end_color="C6EFCE"))
    # grade colours
    grng = f"{grade_col}3:{grade_col}{2+N_STUDENTS}"
    for g, col in [("A*", "C6EFCE"), ("A", "D9EAD3"), ("B", "E2EFDA"),
                   ("C", "FFF2CC"), ("D", "FCE5CD"), ("E", "F8CBAD"), ("U", "F4C7C3")]:
        ws.conditional_formatting.add(grng, CellIsRule(operator="equal", formula=[f'"{g}"'],
            fill=PatternFill("solid", fgColor=col)))
    # class overview rows
    base = 4 + N_STUDENTS
    ws.cell(row=base, column=1, value="CLASS AVERAGE %").font = Font(bold=True)
    for j in range(len(ASSESSMENTS)):
        col = get_column_letter(2 + j)
        ws.cell(row=base, column=2 + j,
                value=f'=IFERROR(AVERAGE({col}3:{col}{2+N_STUDENTS})/{col}2,"")').number_format = "0%"
        ws.cell(row=base, column=2 + j).alignment = CENTER
    ws.cell(row=base, column=2 + len(ASSESSMENTS),
            value=f'=IFERROR(AVERAGE({avg_col}3:{avg_col}{2+N_STUDENTS}),"")').number_format = "0%"
    ws.cell(row=base, column=2 + len(ASSESSMENTS)).alignment = CENTER
    # instructions
    note = base + 2
    for k, txt in enumerate([
        "How to use: enter each assessment's MAX MARK in the yellow row 2, then type each student's raw mark.",
        "Avg % = total marks gained / total marks available across attempted assessments. Grade auto-looks-up the boundaries.",
        "Edit grade boundaries on the 'Grade Boundaries' tab. Replace 'Student N' with real names. Add rows for more students.",
        "KA = Key Assessment (formative), CAP = attainment checkpoint — both from the scheme of work.",
    ]):
        ws.cell(row=note + k, column=1, value=txt).font = Font(italic=True, color="666666")
    return ws


def main():
    wb = openpyxl.Workbook()
    markbook_sheet(wb)
    grade_sheet(wb)
    os.makedirs(OUT, exist_ok=True)
    wb.save(os.path.join(OUT, "H446-Markbook.xlsx"))
    print("wrote Teacher-Toolkit/Markbook/H446-Markbook.xlsx")


if __name__ == "__main__":
    main()
