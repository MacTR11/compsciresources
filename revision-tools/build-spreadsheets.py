#!/usr/bin/env python3
"""
Build interactive .xlsx spreadsheets:
  - Flashcards.xlsx                 (one sheet per topic: Front | Back)
  - Self-Assessment-Tracker.xlsx    (RAG dropdowns + colour conditional formatting)
  - Scheme-of-Work-Homework-and-Consolidation.xlsx (the weekly plan)

Output: Spreadsheets/ at the repo root.
Usage:  python3 revision-tools/build-spreadsheets.py
"""
import os, glob, csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "Spreadsheets")

HEAD_FILL = PatternFill("solid", fgColor="2B6CB0")
HEAD_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="CCCCCC")] * 4)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def topic_title(key):
    import re
    m = re.match(r"^(\d\.\d)-(.*)$", key)
    num = m.group(1) if m else ""
    rest = (m.group(2) if m else key).replace("-", " ").title()
    return f"{num} {rest}".strip()


def build_flashcards():
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for f in sorted(glob.glob(os.path.join(REPO, "revision-tools", "flashcards", "*.csv"))):
        key = os.path.splitext(os.path.basename(f))[0]
        title = key.split("-")[0]  # sheet name = "1.1" etc (<=31 chars, unique)
        ws = wb.create_sheet(title=title)
        ws.append(["Front (question / term)", "Back (answer)"])
        rows = [r for r in csv.reader(open(f, newline="", encoding="utf-8")) if len(r) >= 2 and r[0].strip()]
        for front, back in [(r[0], r[1]) for r in rows]:
            ws.append([front, back])
        ws.column_dimensions["A"].width = 48
        ws.column_dimensions["B"].width = 70
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = WRAP; cell.border = THIN
        style_header(ws, 2)
        ws.cell(row=1, column=1).comment = None
    os.makedirs(OUT, exist_ok=True)
    wb.save(os.path.join(OUT, "Flashcards.xlsx"))
    print("  Flashcards.xlsx")


# (spec, sub-topic) rows for the tracker
TRACKER_ROWS = [
    ("1.1.1", "Structure & function of the processor (ALU, CU, registers, FDE, performance)"),
    ("1.1.2", "Types of processor (Von Neumann/Harvard, CISC/RISC, GPUs, parallel)"),
    ("1.1.3", "Input, output & storage devices; RAM/ROM; virtual storage"),
    ("1.2.1", "Operating systems (memory mgmt, scheduling, interrupts, BIOS, VMs)"),
    ("1.2.2", "Applications generation (translators, stages of compilation, linkers)"),
    ("1.2.3", "Software development (methodologies, lifecycle)"),
    ("1.2.4", "Types of programming language (paradigms, assembly/LMC, addressing, OOP)"),
    ("1.3.1", "Compression, encryption & hashing"),
    ("1.3.2", "Databases (keys, normalisation to 3NF, SQL, ACID)"),
    ("1.3.3", "Networks (LAN/WAN, TCP/IP, protocols, hardware)"),
    ("1.3.4", "Web technologies (HTML/CSS/JS, PageRank, client/server-side)"),
    ("1.4.1", "Data types (binary/hex, two's complement, floating point, shifts)"),
    ("1.4.2", "Data structures (arrays, lists, stacks, queues, trees, graphs, hash tables)"),
    ("1.4.3", "Boolean algebra (gates, De Morgan's, K-maps, adders, flip-flops)"),
    ("1.5.1", "Computing related legislation (DPA, CMA, CDPA, RIPA)"),
    ("1.5.2", "Moral & ethical issues"),
    ("2.1.1", "Thinking abstractly"),
    ("2.1.2", "Thinking ahead (inputs/outputs, preconditions, caching, reuse)"),
    ("2.1.3", "Thinking procedurally (decomposition)"),
    ("2.1.4", "Thinking logically (decision points)"),
    ("2.1.5", "Thinking concurrently"),
    ("2.2.1", "Programming techniques (recursion, scope, parameters, OOP)"),
    ("2.2.2", "Computational methods (backtracking, heuristics, etc.)"),
    ("2.3.1", "Algorithms (Big O, searching, sorting, traversals, Dijkstra/A*)"),
]


def build_tracker():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "RAG Tracker"
    ws.append(["Spec", "Sub-topic", "RAG", "Date reviewed", "Score %", "Notes / what to fix"])
    for spec, name in TRACKER_ROWS:
        ws.append([spec, name, "", "", "", ""])
    widths = {"A": 8, "B": 62, "C": 7, "D": 14, "E": 9, "F": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    style_header(ws, 6)
    last = len(TRACKER_ROWS) + 1
    # RAG dropdown
    dv = DataValidation(type="list", formula1='"R,A,G"', allow_blank=True)
    dv.prompt = "R = can't do it, A = shaky, G = secure"
    ws.add_data_validation(dv); dv.add(f"C2:C{last}")
    # conditional colour fills for R/A/G
    rng = f"C2:C{last}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"R"'],
        fill=PatternFill("solid", fgColor="F4CCCC"), font=Font(bold=True, color="990000")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"A"'],
        fill=PatternFill("solid", fgColor="FCE5CD"), font=Font(bold=True, color="996600")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"G"'],
        fill=PatternFill("solid", fgColor="D9EAD3"), font=Font(bold=True, color="006600")))
    for row in ws.iter_rows(min_row=2, max_row=last):
        for cell in row:
            cell.border = THIN; cell.alignment = WRAP
        row[2].alignment = Alignment(horizontal="center", vertical="center")
    # live counts
    ws.cell(row=last + 2, column=2, value="🔴 Red:").font = Font(bold=True)
    ws.cell(row=last + 2, column=3, value=f'=COUNTIF({rng},"R")')
    ws.cell(row=last + 3, column=2, value="🟠 Amber:").font = Font(bold=True)
    ws.cell(row=last + 3, column=3, value=f'=COUNTIF({rng},"A")')
    ws.cell(row=last + 4, column=2, value="🟢 Green:").font = Font(bold=True)
    ws.cell(row=last + 4, column=3, value=f'=COUNTIF({rng},"G")')
    os.makedirs(OUT, exist_ok=True)
    wb.save(os.path.join(OUT, "Self-Assessment-Tracker.xlsx"))
    print("  Self-Assessment-Tracker.xlsx")


def build_sow():
    src = os.path.join(REPO, "revision-tools", "scheme-of-work",
                       "year-12-homework-and-consolidation.csv")
    if not os.path.exists(src):
        return
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Year 12 HW & Consolidation"
    rows = list(csv.reader(open(src, newline="", encoding="utf-8")))
    for r in rows:
        ws.append(r)
    widths = {"A": 6, "B": 14, "C": 40, "D": 55, "E": 55}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    style_header(ws, 5)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP; cell.border = THIN
        # tint holiday/half-term rows
        if str(row[0].value) in ("HT", "Hol"):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="EFEFEF")
    os.makedirs(OUT, exist_ok=True)
    wb.save(os.path.join(OUT, "Scheme-of-Work-Homework-and-Consolidation.xlsx"))
    print("  Scheme-of-Work-Homework-and-Consolidation.xlsx")


if __name__ == "__main__":
    build_flashcards()
    build_tracker()
    build_sow()
    print("\nDone. Spreadsheets in Spreadsheets/")
